import flet as ft
import openpyxl
import os
import datetime
import tempfile
import platform


def main(page: ft.Page):
    # --- 基础配置 ---
    page.title = "提货明细生成器"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.scroll = ft.ScrollMode.AUTO
    page.padding = 20

    def get_asset_path(filename):
        return os.path.join("assets", filename)

    def show_toast(text):
        sb = ft.SnackBar(ft.Text(text))
        page.overlay.append(sb)
        sb.open = True
        page.update()

    def search_customer(keyword):
        data_path = get_asset_path("data.xlsx")
        if not os.path.exists(data_path):
            show_toast(f"找不到数据源: assets/data.xlsx")
            return []
        try:
            wb = openpyxl.load_workbook(data_path, data_only=True)
            ws = wb["Sheet2"]
            matches = []
            for row in range(1, ws.max_row + 1):
                cell_val = ws.cell(row=row, column=2).value
                if cell_val and keyword in str(cell_val):
                    matches.append({
                        "name": cell_val,
                        "phone": ws.cell(row=row + 1, column=2).value,
                        "addr": ws.cell(row=row + 2, column=2).value,
                        "extra": ws.cell(row=row + 3, column=2).value
                    })
            wb.close()
            return matches
        except Exception as e:
            show_toast(f"读取 Excel 出错: {e}")
            return []

    def generate_and_share(customer_info):
        try:
            tpl_path = get_asset_path("template.xlsx")
            if not os.path.exists(tpl_path):
                show_toast("找不到模板文件 assets/template.xlsx")
                return

            wb = openpyxl.load_workbook(tpl_path)
            ws = wb["1"]

            # 填充数据
            ws["C2"] = datetime.datetime.now().strftime("%Y年%m月%d日")
            ws["B6"] = customer_info["name"]
            ws["E6"] = customer_info["phone"]
            ws["C6"] = customer_info["addr"]
            ws["D6"] = customer_info["extra"]
            ws["G6"] = product_input.value
            ws["J6"] = count_input.value
            ws["M6"] = temp_radio.value

            # 生成路径
            temp_dir = tempfile.gettempdir()
            save_name = f"提货单_{customer_info['name']}.xlsx"
            save_path = os.path.abspath(os.path.join(temp_dir, save_name))

            wb.save(save_path)
            wb.close()

            # --- 分享/打开逻辑 (基于官网不同平台支持度说明) ---
            # 官方文档：share_files 仅支持 Android, iOS, macOS

            system = platform.system().lower()

            # 逻辑：如果是移动端，尝试分享；如果是 Windows，直接打开文件
            if page.platform in ["android", "ios"] or system == "darwin":
                if hasattr(page, "share_files"):
                    page.share_files([save_path])
                    show_toast("已调起系统分享")
                else:
                    show_toast("当前版本不支持 share_files 接口")
            elif system == "windows":
                # Windows 不支持 share_files，我们直接打开生成的 Excel 文件
                show_toast("Windows 不支持分享菜单，已为你直接打开文件")
                os.startfile(save_path)
            else:
                show_toast(f"存至: {save_path}")

        except Exception as e:
            show_toast(f"生成失败: {e}")

    def handle_gen_click(e):
        if not search_input.value:
            show_toast("请输入搜索关键字")
            return

        results = search_customer(search_input.value)
        if not results:
            show_toast("未找到匹配客户")
            return

        if len(results) > 1:
            options = []
            for item in results:
                # 使用函数闭包，并在点击后关闭 BottomSheet
                def on_select(info):
                    return lambda _: [
                        setattr(bottom_sheet, "open", False),  # 关闭面板
                        page.update(),
                        generate_and_share(info)
                    ]

                options.append(ft.ListTile(
                    title=ft.Text(item["name"]),
                    subtitle=ft.Text(f"{item['addr'] or ''}"),
                    on_click=on_select(item)
                ))

            bottom_sheet.content = ft.Container(
                content=ft.Column(options, tight=True, scroll=ft.ScrollMode.AUTO),
                padding=10,
                height=400
            )
            # --- 兼容旧版语法：设置属性并 update ---
            bottom_sheet.open = True
            page.update()
        else:
            generate_and_share(results[0])

    # --- UI 组件 ---
    search_input = ft.TextField(label="🔍 搜索客户", border_radius=12)
    product_input = ft.TextField(label="📦 产品名称", border_radius=12)
    count_input = ft.TextField(label="📊 件数", keyboard_type=ft.KeyboardType.NUMBER, border_radius=12)

    temp_radio = ft.RadioGroup(
        content=ft.Row([
            ft.Radio(value="常温", label="常温"),
            ft.Radio(value="冷链", label="冷链")
        ], alignment=ft.MainAxisAlignment.CENTER),
        value="常温"
    )

    # 底部面板
    bottom_sheet = ft.BottomSheet(ft.Container(padding=10))
    page.overlay.append(bottom_sheet)

    page.add(
        ft.Container(height=10),
        ft.Text("🦅 提货明细生成器", size=26, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_700),
        ft.Divider(height=20),
        search_input,
        product_input,
        count_input,
        ft.Row([ft.Text("🌡️ 温度选择:"), temp_radio], alignment=ft.MainAxisAlignment.CENTER),
        ft.Container(height=20),
        ft.ElevatedButton(
            content=ft.Row([ft.Icon(ft.Icons.SEND), ft.Text("生成并分享文件")], alignment=ft.MainAxisAlignment.CENTER),
            width=300, height=50, on_click=handle_gen_click,
            bgcolor=ft.Colors.BLUE_600, color=ft.Colors.WHITE
        )
    )


if __name__ == "__main__":
    ft.app(target=main, assets_dir="assets")
